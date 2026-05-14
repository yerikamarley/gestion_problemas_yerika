from __future__ import annotations

import json
import os
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell


PROJECT_DIR = Path(__file__).resolve().parents[1]
DOWNLOADS_DIR = Path.home() / "Downloads"
INPUT = Path(
    os.environ.get(
        "INCIDENTES_XLSX",
        DOWNLOADS_DIR / "CNC_MST-PR-002 Gestión de incidentes_v6.xlsx",
    )
)
OUT_DIR = PROJECT_DIR / "outputs" / "analisis_incidentes"
OUT_DIR.mkdir(parents=True, exist_ok=True)


def clean(value):
    if value is None:
        return None
    if hasattr(value, "isoformat"):
        return value.isoformat()
    text = str(value).strip()
    return text if text else None


def row_values(ws, row_idx, max_col):
    values = []
    for col_idx in range(1, max_col + 1):
        cell = ws.cell(row_idx, col_idx)
        values.append(clean(cell.value))
    while values and values[-1] is None:
        values.pop()
    return values


def non_empty_rows(ws, limit=80):
    rows = []
    max_col = min(ws.max_column, 30)
    for row_idx in range(1, ws.max_row + 1):
        values = row_values(ws, row_idx, max_col)
        if any(v is not None for v in values):
            rows.append({"row": row_idx, "values": values})
            if len(rows) >= limit:
                break
    return rows


def formula_count(ws):
    count = 0
    examples = []
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            if isinstance(cell.value, str) and cell.value.startswith("="):
                count += 1
                if len(examples) < 20:
                    examples.append({"cell": cell.coordinate, "formula": cell.value})
    return count, examples


def style_fill_count(ws):
    fills = Counter()
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            fill = cell.fill
            color = getattr(fill.fgColor, "rgb", None) or getattr(fill.fgColor, "indexed", None)
            if fill.fill_type and color:
                fills[str(color)] += 1
    return fills.most_common(10)


workbook = load_workbook(INPUT, data_only=False, read_only=False)
summary = {
    "file": str(INPUT),
    "sheet_count": len(workbook.worksheets),
    "sheets": [],
}

for ws in workbook.worksheets:
    formula_total, formulas = formula_count(ws)
    sheet_info = {
        "title": ws.title,
        "sheet_state": ws.sheet_state,
        "max_row": ws.max_row,
        "max_column": ws.max_column,
        "merged_ranges": [str(rng) for rng in list(ws.merged_cells.ranges)[:30]],
        "tables": list(ws.tables.keys()),
        "auto_filter": str(ws.auto_filter.ref) if ws.auto_filter and ws.auto_filter.ref else None,
        "freeze_panes": str(ws.freeze_panes) if ws.freeze_panes else None,
        "formula_count": formula_total,
        "formula_examples": formulas,
        "top_fills": style_fill_count(ws),
        "first_non_empty_rows": non_empty_rows(ws, limit=100),
    }
    summary["sheets"].append(sheet_info)

(OUT_DIR / "openpyxl_summary.json").write_text(
    json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8"
)

print(json.dumps({
    "file": str(INPUT),
    "sheet_count": summary["sheet_count"],
    "sheets": [
        {
            "title": s["title"],
            "state": s["sheet_state"],
            "rows": s["max_row"],
            "columns": s["max_column"],
            "formulas": s["formula_count"],
            "tables": s["tables"],
        }
        for s in summary["sheets"]
    ],
    "output": str(OUT_DIR / "openpyxl_summary.json"),
}, ensure_ascii=False, indent=2))
