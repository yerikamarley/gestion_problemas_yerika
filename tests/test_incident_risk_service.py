import unittest

import pandas as pd
from openpyxl import load_workbook

from services.incident_risk_service import build_analysis, classify_incident, classify_incidents, group_materialized_events
from utils.risk_excel_export import build_risk_workbook


class IncidentRiskServiceTest(unittest.TestCase):
    def classify(self, **values):
        values.setdefault("numero", "INC0001")
        return classify_incident(values)

    def test_r76(self):
        result = self.classify(servicio_negocio="Biometría", descripcion="Falla de validación de identidad con código OTP")
        self.assertEqual(("RISK", "R76"), (result["classification_type"], result["risk_id"]))

    def test_false_alarm(self):
        result = self.classify(descripcion="Alerta tipificada como falsa alarma sin afectación")
        self.assertEqual(("EXCLUSION", "Falsas Alarmas (Monitoreo NOC/SOC)"), (result["classification_type"], result["exclusion_category"]))

    def test_ec2_504_integration_is_r164_not_r140(self):
        result = self.classify(servicio_negocio="EC2", descripcion="Error 504 integración consulta de listas")
        self.assertEqual("R164", result["risk_id"]); self.assertNotIn("R140", result["risk_id"])

    def test_real_ocsp_is_r140(self):
        result = self.classify(servicio_negocio="OCSP", descripcion="Caída real, OCSP no responde")
        self.assertEqual("R140", result["risk_id"])

    def test_password_is_bau(self):
        result = self.classify(descripcion="Olvido de contraseña")
        self.assertEqual(("EXCLUSION", "Requerimientos y Soporte Rutinario (BAU)"), (result["classification_type"], result["exclusion_category"]))

    def test_installation_without_failure_is_bau(self):
        result = self.classify(descripcion="Solicitud de instalación y activación de Certitoken en nuevo PC")
        self.assertEqual(("EXCLUSION", "Requerimientos y Soporte Rutinario (BAU)"), (result["classification_type"], result["exclusion_category"]))

    def test_rpost_and_certitoken_downtime_are_infrastructure(self):
        self.assertEqual("R140", self.classify(servicio_negocio="RPOST", descripcion="Portal RPOST caído y no responde")["risk_id"])
        self.assertEqual("R140", self.classify(servicio_negocio="Certitoken", descripcion="Certitoken caído e indisponible")["risk_id"])

    def test_monitoring_failure_has_specific_risk(self):
        result = self.classify(servicio_negocio="NOC", descripcion="El monitoreo no generó alerta durante la caída")
        self.assertEqual("R-MON", result["risk_id"])

    def test_ambiguous_is_pending(self):
        self.assertEqual("PENDING", self.classify(descripcion="Se reporta una novedad")["classification_type"])

    def test_overlap_has_single_priority_result(self):
        result = self.classify(servicio_negocio="OCSP EC2", descripcion="Error 504 integración, caída OCSP")
        self.assertEqual("R164", result["risk_id"])

    def test_reconciliation_counts_unique_incidents(self):
        df = pd.DataFrame([
            {"numero":"INC1", "creado":"2026-01-01", "servicio_negocio":"Biometría", "descripcion":"Falla OTP validación de identidad"},
            {"numero":"INC2", "creado":"2026-01-02", "descripcion":"Olvido de contraseña"},
            {"numero":"INC3", "creado":"2026-01-03", "descripcion":"Novedad no concluyente"},
        ])
        analysis = build_analysis(classify_incidents(df), [1])
        v = analysis["validation"]
        self.assertTrue(v["reconciled"]); self.assertEqual(v["total"], v["risk"] + v["exclusion"] + v["pending"])
        self.assertTrue(all(analysis["risks"]["Cantidad tickets asociados"] == analysis["risks"]["Tickets Asociados"].str.split(", ").str.len()))

    def test_excel_has_required_sheets(self):
        df = pd.DataFrame([
            {"numero":"INC1", "creado":"2026-01-01", "servicio_negocio":"OCSP", "descripcion":"Caída OCSP no responde", "tipificacion_auto":"Indisponibilidad", "causa_raiz_auto":"Agotamiento de conexiones"},
            {"numero":"INC2", "creado":"2026-01-03", "servicio_negocio":"OCSP", "descripcion":"Nueva caída OCSP no responde", "tipificacion_auto":"Indisponibilidad", "causa_raiz_auto":"Agotamiento de conexiones"},
        ])
        analysis = build_analysis(classify_incidents(df), [1])
        from io import BytesIO
        wb = load_workbook(BytesIO(build_risk_workbook(analysis, 2026, 1, 1)))
        required = {"Informe Ejecutivo", "Resumen", "Riesgos", "Conciliación", "Exclusiones", "Metodología"}
        self.assertEqual(required, set(wb.sheetnames))
        self.assertEqual("Matriz ejecutiva de riesgos materializados", wb["Informe Ejecutivo"]["A1"].value)
        self.assertEqual("INC asociados", wb["Informe Ejecutivo"]["F8"].value)
        self.assertEqual("Nivel de recurrencia", wb["Informe Ejecutivo"]["G8"].value)
        self.assertIn("INC1", wb["Informe Ejecutivo"]["F9"].value)
        self.assertEqual("REINCIDENTE", wb["Informe Ejecutivo"]["G9"].value)
        self.assertEqual("Indisponibilidad", wb["Informe Ejecutivo"]["C9"].value)
        self.assertEqual("Agotamiento de conexiones", wb["Informe Ejecutivo"]["D9"].value)
        self.assertEqual(required, set(wb.sheetnames))
        for sheet in wb.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        self.assertEqual("Arial", cell.font.name)
                        self.assertEqual(12, cell.font.sz)
                        self.assertEqual("000000", cell.font.color.rgb[-6:])

    def test_groups_multiple_tickets_into_one_event_and_separates_recurrence(self):
        df = pd.DataFrame([
            {"numero":"INC1", "creado":"2026-01-01 08:00", "servicio_negocio":"OCSP", "descripcion":"Caída OCSP no responde"},
            {"numero":"INC2", "creado":"2026-01-01 09:30", "servicio_negocio":"OCSP", "descripcion":"OCSP continúa caído"},
            {"numero":"INC3", "creado":"2026-01-03 08:00", "servicio_negocio":"OCSP", "descripcion":"Nueva caída OCSP no responde"},
        ])
        events = group_materialized_events(classify_incidents(df))
        self.assertEqual(2, len(events))
        self.assertEqual(2, int(events.iloc[0]["ticket_count"]))
        self.assertTrue((events["event_status"] == "REINCIDENTE").all())


if __name__ == "__main__": unittest.main()
