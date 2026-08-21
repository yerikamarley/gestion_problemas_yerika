"""Generación del informe ejecutivo anonimizado de riesgos materializados."""

from io import BytesIO

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

PURPLE, YELLOW, ORANGE, WHITE, BLACK = "C9B2E6", "FFE699", "F4B183", "FFFFFF", "000000"
THIN = Side(style="thin", color="A6A6A6")
MONTHS = {1: "Ene", 2: "Feb", 3: "Mar", 4: "Abr", 5: "May", 6: "Jun", 7: "Jul", 8: "Ago", 9: "Sep", 10: "Oct", 11: "Nov", 12: "Dic"}


def _safe(value):
    if value is None or (not isinstance(value, (list, tuple, dict)) and pd.isna(value)):
        return None
    return value.to_pydatetime().replace(tzinfo=None) if isinstance(value, pd.Timestamp) else value


def _font(*, bold=False, italic=False):
    return Font(name="Arial", size=12, bold=bold, italic=italic, color=BLACK)


def _header(ws, row, start, end, fill=PURPLE):
    for col in range(start, end + 1):
        cell = ws.cell(row, col)
        cell.font = _font(bold=True)
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _body(ws, first, last, start, end):
    for row in ws.iter_rows(min_row=first, max_row=max(first, last), min_col=start, max_col=end):
        for cell in row:
            cell.font = _font()
            cell.fill = PatternFill("solid", fgColor=WHITE)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _section(cell, text):
    cell.value = text
    cell.font = _font(bold=True)
    cell.fill = PatternFill("solid", fgColor=YELLOW)
    cell.alignment = Alignment(vertical="center", wrap_text=True)


def _write_df(ws, df, header_fill=PURPLE):
    columns = list(df.columns)
    ws.append(columns)
    for row in df.itertuples(index=False, name=None):
        ws.append([_safe(value) for value in row])
    if columns:
        _header(ws, 1, 1, len(columns), header_fill)
        _body(ws, 2, ws.max_row, 1, len(columns))
        ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{max(1, ws.max_row)}"
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False
    for index, column in enumerate(columns, 1):
        values = [str(column)] + [str(value or "") for value in df[column].head(200)]
        ws.column_dimensions[get_column_letter(index)].width = min(48, max(13, max(map(len, values)) + 2))


def _treatment(row):
    has_problem = bool(str(row.get("Problemas asociados") or "").strip())
    state = row.get("Estado del problema") or "Sin tratamiento formal asociado"
    cause = row.get("Causa raíz consolidada") or "En proceso de análisis"
    action = "Seguimiento mediante un plan de tratamiento relacionado." if has_problem else "Evaluar la creación de un plan de tratamiento y acciones preventivas."
    cause_status = row.get("Estado causa raíz") or "Pendiente de investigación"
    return f"Causa raíz: {cause}\nEstado de la causa: {cause_status}\nEstado del problema: {state}\nAcción: {action}"


def _public_risks(risks):
    columns = ["ID", "Riesgo Materializado", "Naturaleza consolidada de los INC", "Causa raíz consolidada", "Estado causa raíz", "Dueño del Riesgo", "Impacto Escala", "Cantidad tickets asociados", "Tickets Asociados", "Eventos reales", "Estado", "Asignación Operativa RACI"]
    result = risks[[column for column in columns if column in risks]].copy()
    if "Estado" in result:
        result["Estado"] = result["Estado"].astype(str).str.replace("🔥 ", "", regex=False)
    return result.rename(columns={
        "Cantidad tickets asociados": "Cantidad de casos asociados",
        "Tickets Asociados": "INC asociados",
        "Estado": "Nivel de recurrencia",
    })


def _public_exclusions(exclusions):
    """Conserva la metodología y elimina ejemplos que puedan identificar casos."""
    result = exclusions.copy()
    for column in result.columns:
        if "justific" in str(column).casefold() or "ejemplo" in str(column).casefold():
            result[column] = result[column].fillna("").astype(str).str.split("Ejemplos:", n=1).str[0].str.strip()
            result = result.rename(columns={column: "Justificación metodológica"})
    return result


def _public_patterns(patterns):
    columns = [
        "id_riesgo", "riesgo_materializado", "criterio_similitud", "dominio_evento", "proveedor_evento",
        "naturaleza_evento", "causa_probable", "estado_causa", "cantidad_incidentes",
        "incidentes_asociados", "impacto_escala", "problemas_plan_trabajo",
    ]
    result = patterns.reindex(columns=columns).copy()
    return result.rename(columns={
        "id_riesgo": "ID riesgo", "riesgo_materializado": "Riesgo materializado",
        "criterio_similitud": "Servicio o componente", "dominio_evento": "Dominio",
        "proveedor_evento": "Proveedor", "naturaleza_evento": "Naturaleza",
        "causa_probable": "Causa probable", "estado_causa": "Estado de la causa",
        "cantidad_incidentes": "Cantidad de INC", "incidentes_asociados": "INC asociados",
        "impacto_escala": "Nivel de recurrencia", "problemas_plan_trabajo": "Problemas asociados",
    })


def _executive(wb, analysis, year, month_from, month_to):
    ws = wb.create_sheet("Informe Ejecutivo")
    ws.sheet_view.showGridLines = False
    period = f"Año {year}" if (month_from, month_to) == (1, 12) else f"{year} · {MONTHS[month_from]}-{MONTHS[month_to]}"
    ws.merge_cells("A1:L1")
    ws["A1"] = "Matriz ejecutiva de riesgos materializados por servicio o componente"
    ws["A1"].font = _font(bold=True)
    ws["A1"].fill = PatternFill("solid", fgColor=PURPLE)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    validation = analysis["validation"]
    ws.merge_cells("A2:L2")
    ws["A2"] = f"Periodo: {period} | {validation['total']} casos únicos analizados"
    ws["A2"].font = _font(italic=True)
    ws["A2"].fill = PatternFill("solid", fgColor=YELLOW)
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("A3:L3")
    ws["A3"] = "Trazabilidad controlada: se incluyen los números de INC asociados, sin notas, solicitudes, datos personales ni detalles técnicos."
    ws["A3"].font = _font(bold=True)
    ws["A3"].fill = PatternFill("solid", fgColor=ORANGE)
    ws["A3"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.merge_cells("A5:L5")
    _section(ws["A5"], "Tabla 1. Riesgos materializados consolidados por servicio o componente")
    risk_count = int(analysis["risks"]["ID"].nunique()) if not analysis["risks"].empty else 0
    ws["A6"] = f"Total de riesgos materializados en el periodo: {risk_count}"
    ws["A6"].font = _font(bold=True)
    headers = ["Servicio o componente", "Riesgo relacionado", "Naturaleza de los INC", "Causa raíz consolidada", "Estado de la causa", "INC asociados", "Nivel de recurrencia", "Cantidad de INC", "Proveedor", "Dominio", "Problemas asociados", "ID del riesgo"]
    for col, value in enumerate(headers, 1):
        ws.cell(8, col, value)
    _header(ws, 8, 1, 12)

    row_no = 9
    patterns = analysis.get("patterns", pd.DataFrame())
    for _, row in patterns.iterrows():
        values = [row.get("criterio_similitud"), row.get("riesgo_materializado"), row.get("naturaleza_evento"), row.get("causa_probable"), row.get("estado_causa"), row.get("incidentes_asociados"), row.get("impacto_escala"), row.get("cantidad_incidentes"), row.get("proveedor_evento"), row.get("dominio_evento"), row.get("problemas_plan_trabajo"), row.get("id_riesgo")]
        for col, value in enumerate(values, 1):
            ws.cell(row_no, col, _safe(value))
        ws.row_dimensions[row_no].height = 105
        row_no += 1
    _body(ws, 9, row_no - 1, 1, 12)
    for current_row in range(9, row_no):
        if str(ws.cell(current_row, 7).value).upper() in {"ALTA", "REINCIDENTE"}:
            ws.cell(current_row, 7).fill = PatternFill("solid", fgColor=ORANGE)
            ws.cell(current_row, 7).font = _font(bold=True)

    row_no += 2
    ws.merge_cells(start_row=row_no, start_column=1, end_row=row_no, end_column=3)
    _section(ws.cell(row_no, 1), "Tabla 2. Conciliación de casos")
    row_no += 1
    for col, value in enumerate(["Categoría", "Cantidad", "Impacto en la matriz"], 1):
        ws.cell(row_no, col, value)
    _header(ws, row_no, 1, 3, fill=YELLOW)
    row_no += 1
    rows = [
        (f"Riesgos materializados ({risk_count} riesgos)", validation["risk"], "Sí"),
        ("Falsos positivos y exclusiones", validation["exclusion"], "No"),
        ("Pendientes de clasificación", validation["pending"], "En análisis"),
        ("Total de casos registrados", validation["total"], "100% conciliado" if validation["reconciled"] else "Requiere revisión"),
    ]
    start = row_no
    for values in rows:
        for col, value in enumerate(values, 1):
            ws.cell(row_no, col, value)
        row_no += 1
    _body(ws, start, row_no - 1, 1, 3)
    if not validation["reconciled"]:
        for cell in ws[row_no - 1][:3]:
            cell.fill = PatternFill("solid", fgColor=ORANGE)

    for index, width in enumerate([25, 42, 48, 48, 23, 38, 20, 17, 22, 22, 35, 15], 1):
        ws.column_dimensions[get_column_letter(index)].width = width
    ws.freeze_panes = "A9"
    ws.auto_filter.ref = f"A8:L{max(8, 8 + len(patterns))}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True


def _trend_sheet(wb, summary, monthly):
    """Crea una vista gráfica y tabular de recurrencia mensual por servicio."""
    ws = wb.create_sheet("Tendencia Servicios")
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Reincidencias mensuales por servicio o componente"
    ws["A1"].font = _font(bold=True)
    ws["A1"].fill = PatternFill("solid", fgColor=PURPLE)
    ws.merge_cells("A1:L1")
    ws["A2"] = "Cada servicio aparece una sola vez; la tabla inferior muestra su recurrencia y los INC que sustentan el consolidado."
    ws["A2"].fill = PatternFill("solid", fgColor=YELLOW)
    ws.merge_cells("A2:L2")
    if monthly is None or monthly.empty:
        ws["A4"] = "No hay datos mensuales disponibles para el periodo seleccionado."
        return
    pivot = monthly.pivot_table(index="mes", columns="causa", values="incidentes", aggfunc="sum", fill_value=0).sort_index()
    hidden_col = 14
    ws.cell(1, hidden_col, "Mes")
    for col, component in enumerate(pivot.columns, hidden_col + 1):
        ws.cell(1, col, component)
    for excel_row, (month, values) in enumerate(pivot.iterrows(), 2):
        ws.cell(excel_row, hidden_col, month)
        for col, value in enumerate(values, hidden_col + 1):
            ws.cell(excel_row, col, int(value))
    chart = LineChart()
    chart.title = "Reincidencias mensuales por servicio o componente"
    chart.y_axis.title = "Incidentes"
    chart.x_axis.title = "Mes"
    chart.height, chart.width = 11, 24
    chart.add_data(Reference(ws, min_col=hidden_col + 1, max_col=hidden_col + len(pivot.columns), min_row=1, max_row=1 + len(pivot)), titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=hidden_col, min_row=2, max_row=1 + len(pivot)))
    for index, series in enumerate(chart.series):
        series.graphicalProperties.line.solidFill = [PURPLE, ORANGE, YELLOW][index % 3]
        series.graphicalProperties.line.width = 22000
    ws.add_chart(chart, "A4")
    table_row = 26
    public = summary.rename(columns={"causa": "Servicio o componente", "total_anual": "Total del periodo", "meses_con_eventos": "Meses con eventos", "primer_incidente": "Primer incidente", "ultimo_incidente": "Último incidente", "incidentes_asociados": "INC asociados", "nivel": "Nivel de recurrencia", "nivel_reincidencia": "Nivel de recurrencia"})
    for col, value in enumerate(public.columns, 1):
        ws.cell(table_row, col, value)
    _header(ws, table_row, 1, len(public.columns))
    for row in public.itertuples(index=False, name=None):
        table_row += 1
        for col, value in enumerate(row, 1):
            ws.cell(table_row, col, _safe(value))
    _body(ws, 27, table_row, 1, len(public.columns))
    ws.freeze_panes = "A27"
    for index, width in enumerate([28, 18, 20, 18, 18, 48, 20], 1):
        ws.column_dimensions[get_column_letter(index)].width = width
    for col in range(hidden_col, hidden_col + len(pivot.columns) + 1):
        ws.column_dimensions[get_column_letter(col)].hidden = True


def build_risk_workbook(analysis, year, month_from, month_to):
    """Devuelve un XLSX ejecutivo sin campos identificables ni notas operativas."""
    wb = Workbook()
    wb.remove(wb.active)
    _executive(wb, analysis, year, month_from, month_to)
    _trend_sheet(wb, analysis.get("pattern_summary", pd.DataFrame()), analysis.get("pattern_monthly", pd.DataFrame()))

    validation = analysis["validation"]
    top = analysis["risks"].iloc[0]["ID"] if not analysis["risks"].empty else "N/A"
    summary = pd.DataFrame([
        ("Periodo analizado", f"{year}-{month_from:02d} a {year}-{month_to:02d}"),
        ("Total de casos únicos", validation["total"]),
        ("Casos materializados", validation["risk"]),
        ("Exclusiones", validation["exclusion"]),
        ("Pendientes", validation["pending"]),
        ("Porcentaje conciliado", validation["percentage"]),
        ("Riesgos materializados", analysis["risks"]["ID"].nunique() if not analysis["risks"].empty else 0),
        ("Eventos materiales estimados", len(analysis.get("events", pd.DataFrame()))),
        ("Riesgo más recurrente", top),
    ], columns=["Indicador", "Valor"])
    methodology = pd.DataFrame([
        ("Alcance", "Informe ejecutivo consolidado y anonimizado."),
        ("Trazabilidad incluida", "Números de los INC asociados a cada riesgo y su nivel de recurrencia."),
        ("Información excluida", "Nombres, notas, solicitudes, descripciones completas, evidencias técnicas sensibles y datos personales."),
        ("Causa pendiente", "Se reporta como 'Causa en proceso de análisis' hasta confirmar la causa raíz."),
        ("Acceso al detalle", "El detalle permanece en el sistema de gestión y requiere autorización según el rol."),
        ("Uso", "Seguimiento directivo, control de tendencias y toma de decisiones."),
    ], columns=["Criterio", "Aplicación"])

    sheets = {
        "Resumen": summary,
        "Patrones Operativos": _public_patterns(analysis.get("patterns", pd.DataFrame())),
        "Riesgos": _public_risks(analysis["risks"]),
        "Conciliación": analysis["reconciliation"],
        "Exclusiones": _public_exclusions(analysis["exclusions"]),
        "Metodología": methodology,
    }
    for name, frame in sheets.items():
        _write_df(wb.create_sheet(name), frame, header_fill=YELLOW if name == "Resumen" else PURPLE)
    wb["Resumen"]["B7"].number_format = "0.00%"

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    cell.font = _font(bold=cell.font.bold, italic=cell.font.italic)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
